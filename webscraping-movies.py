
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

# print(title.text)

table_rows = soup.findAll("tr")

wb = xl.Workbook()
ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = "No."
ws['A1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['A'].width = 6

ws['B1'] = 'Movie Title'
ws['B1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['B'].width = 26

ws['C1'] = "Release Date"
ws['C1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['C'].width = 20

ws['D1'] = "Total Gross"
ws['D1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['D'].width = 20

ws['E1'] = "Theaters"
ws['E1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['E'].width = 20

ws['F1'] = "Average per theater"
ws['F1'].font = Font(name = 'Times New Roman', size=18, italic=False,bold=True)
ws.column_dimensions['F'].width = 28


row_number = 2

for row in table_rows[1:6]:

    columns = row.findAll("td")

    rank = columns[0].text
    titles = columns[1].text
    release_date = columns[8].text
    total_gross = int(columns[7].text.replace('$','').replace(',',''))
    theatres = int(columns[6].text.replace(',',''))
    average = total_gross/theatres


    ws.cell(row=row_number, column = 1, value = rank)
    ws.cell(row=row_number, column = 2, value = titles)
    ws.cell(row=row_number, column = 3, value = release_date)
    ws.cell(row=row_number, column = 4, value = total_gross)
    ws.cell(row=row_number, column = 5, value = theatres)
    ws.cell(row=row_number, column = 6, value = average)

    row_number += 1


for cell in ws["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["F:F"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["E:E"]:
    cell.number_format = '#,##0'


wb.save('BoxOffice Report.xlsx')


